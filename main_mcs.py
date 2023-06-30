# https://istories.media/workshops/2021/09/20/parsing-s-pomoshchyu-python-urok-2/
# https://questu.ru/articles/81673/
# https://stackoverflow.com/questions/29858752/error-message-chromedriver-executable-needs-to-be-available-in-the-path
# https://www.geeksforgeeks.org/python-tkinter-scrolledtext-widget/
import threading
# https://stackoverflow.com/questions/27050492/how-do-you-create-a-tkinter-gui-stop-button-to-break-an-infinite-loop
# https://ru.stackoverflow.com/questions/1194013/%D0%9F%D1%80%D0%B8%D0%BE%D1%81%D1%82%D0%B0%D0%BD%D0%BE%D0%B2%D0%BA%D0%B0-%D0%BF%D1%80%D0%BE%D0%B3%D1%80%D0%B0%D0%BC%D0%BC%D1%8B-%D0%BF%D0%BE-%D0%BD%D0%B0%D0%B6%D0%B0%D1%82%D0%B8%D1%8E-%D0%BA%D0%BD%D0%BE%D0%BF%D0%BA%D0%B8
import tkinter

from selenium.webdriver import Chrome
# from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from tkinter import *
from tkinter.messagebox import showerror, showinfo
import tkinter.ttk as ttk
from tkinter import scrolledtext
from openpyxl import load_workbook
import logging
from bs4 import BeautifulSoup
import lxml
import re
from threading import Thread
from tkinter import filedialog as fd
import analiz_place
import analiz_snils
import analiz_serial_number

logging.basicConfig(level=logging.DEBUG, filename='log.log', filemode='w',
                    format="%(asctime)s %(levelname)s %(message)s")

xpath_login = '//*[@id="login"]'
xpath_password = '//*[@id="password"]'
xpath_button = '/html/body/div/div/div/div/div/div/div/div[3]/form/button'          # кнопка авторизации
xpath_go = '/html/body/div/div/div/div/div/div[4]/div/div/div/div[2]/div/div[2]/div/div/div/a'           # ссылка второй страницы, "перейти"
xpath_notice = '//*[@id="block-system-main-menu"]/ul/li/ul/li[3]/ul/li[4]/a'        # уведомление
xpath_filter_open = '/html/body/div[2]/div/section/h1/span[2]'
xpath_apply = '//*[@id="edit-submit-fgpn-license-notification-record"]'         # кнопка применить
xpath_number = '//*[@id="edit-field-gl-registry-num-value"]'           # ссылка на поле с вводом номера регистрации
xpath_open = '//*[@id="block-system-main"]/div/div[2]/div/table/tbody/tr/td[12]/div/a/span'     # забавный файлик
xpath_open_a = '//*[@id="bootstrap-panel-3-body"]/div[2]/div[2]/div/a'
class_date_registration = 'field-name-field-gl-registry-date'
class_date_end = 'field-name-field-fgpn-notify-end-date'
class_license = 'field-name-field-fs-subject'
class_number_l = 'field-name-field-gl-number'

url_entry = 'https://passport.cgu.mchs.ru/oauth/login?login_challenge=9043dbc4bed146c3ae16ef4e6c39fa7e'
s = Service("chromedriver.exe")
browser = Chrome(service=s)

flag_pause = False
k = 2

def pause_check():
    if flag_pause:
        btn_pause.configure(state='normal')
        btn_pause.configure(text='Продолжить')
        while flag_pause:
            pass

def entry():
    try:
        browser.get(url_entry)
        logging.info(f'успешное открытие страницы регистрации {url_entry}')
    except Exception:
        logging.exception(f'Не удается открыть {url_entry}')
        showerror('Ошибка', 'Не удается перейти на страницу входа')
        return

    try:
        br_login = browser.find_element(By.XPATH, xpath_login)
        br_password = browser.find_element(By.XPATH, xpath_password)
        br_button = browser.find_element(By.XPATH, xpath_button)
        logging.info(f'Элементы br_button, br_login, br_password успешно найдены')
    except Exception:
        showerror('Ошибка', 'Не найден элемент')
        logging.exception(f'Один из элементов br_button, br_login, br_password не найден')
        return
    login = txt_login.get()
    password = txt_password.get()
    br_login.send_keys(login)
    br_password.send_keys(password)
    br_button.click()
    # browser.refresh()
    # sleep(2)
    br_login = browser.find_elements(By.XPATH, xpath_login)
    if len(br_login) == 0:
        showinfo('Вход', 'Успешный вход')
        btn_entry.configure(state='disabled')
        try:
            br_go = browser.find_element(By.XPATH, xpath_go)            # ищем ссылку "перейти"
        except Exception:
            showerror('Ошибка', 'Не найдена ссылка')
            return
        br_go.click()           # нажимаем на ссылку перейти
        sleep(2)
        try:
            browser.switch_to.window(browser.window_handles[1])         # переключаемся на вторую вкладку
        except Exception:
            showerror('Ошибка', 'Ожидание второй вкладки, а в браузере всего лишь одна')
        # try:
        #     br_notice = browser.find_element(By.XPATH, xpath_notice)        # ищем ссылку под названием "уведомление"
        # except Exception:
        #     showerror('Ошибка', 'Не найдена ссылка')
        #     return
        # br_notice.click()           # кликаем и переходим на страницу с поиском
        txt_folder_xl.configure(state='normal')
        btn_start.configure(state='normal')
        btn_check.configure(state='normal')
        sleep(2)
    else:
        showerror('Ошибка', 'Ошибка входа')
        logging.exception('Ошибка входа')


def get_number(folder):
    logging.info('Успешно вошли в функцию get_number, в потоке')
    # # pause_check()
    progress_bar['value'] = 0
    window.update()
    try:
        wb = load_workbook(folder)
        logging.info(f'Книга успешно открыта по адресу {folder}')
    except Exception:
        showerror('Ошибка', 'Не удается подключиться к книге Excel')
        logging.exception(f'Не удается подключиться к книге по пути {folder}')
        return
    try:
        ws = wb['Главный_лист']
        logging.info('Лист Excel успешно открыт')
    except Exception:
        showerror('Ошибка', 'Не найден лист')
        logging.exception(f'Лист не найден')
        return
    try:
        wb.save(folder)
    except PermissionError:
        showerror('Ошибка', f'Необходимо закрыть файл {folder}. Закройте файл и нажмите на кнопку заново заново')
        logging.exception('файл открыт')
        wb.close()
        return
    else:
        logging.info('Файл закрыт, идем дальше')
    try:
        br_filter_open = browser.find_element(By.XPATH, xpath_filter_open)      # находим кнопку фильтра
        logging.info("The filter is opening successfully")
    except Exception:
        showerror('Не найден фильтр')
        logging.exception("The filter isn't opening successfully")
        wb.close()
        return
    br_filter_open.click()          # кликаем на кнопку фильтра
    numbers_excel = ws['C']         # берем все ячейки C
    count = len(numbers_excel)      # считаем кол-во ячеек непустых
    progress_bar['value'] += 1
    window.update()
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3, max_row=count):
        for cell in col:
            # # pause_check()
            if progress_bar['value'] > 100:
                progress_bar['value'] = 0
                window.update()
            number = cell.value
            logging.info(f"Take number {number} from worklist")
            if number is None:          # если в excel пустоя ячейка (отсутствует номер), исключаем None. Иначе ошибка TypeError в send_keys()
                logging.info(f"Take number number from worklist, number is {number}, break")
                break
            try:
                br_number = browser.find_element(By.XPATH, xpath_number)  # находим поле с вводом номера регистрации
                logging.info('Элемент br_number упешно найден')
            except Exception:
                showerror('Ошибка', 'Не удается найти элемент')
                logging.exception('Элемент br_number не найден')
                wb.close()
                return
            # pause_check()
            try:
                br_apply = browser.find_element(By.XPATH, xpath_apply)  # находим поле с кнопкой "применить"
                logging.info('Элемент br_apply упешно найден')
            except Exception:
                showerror('Ошибка', 'Не удается найти элемент')
                logging.exception('Элемент br_apply не найден')
                wb.close()
                return
            logging.info('Clear the input textarea')
            br_number.clear()
            # pause_check()
            try:
                br_number.send_keys(number)         # отправляем номер в поле регистрации
                logging.info("the number from excel is sended in the br_number complitely ")
            except Exception:
                logging.exception("the number from excel is not sended in br_number")
                wb.close()
                return
            try:
                br_apply.click()            # нажимаем на кнопку "применить"
                logging.info("Click on the button_apply is successful")
            except Exception:
                logging.exception("Click on the button_apply isn't successful")
                wb.close()
                return
            sleep(2)
            try:
                br_open = browser.find_element(By.XPATH, xpath_open)           # находим забавный файлик
                logging.info('Элемент br_open найден')
            except Exception:
                # showerror('Ошибка', 'Не найден элемент')
                logging.exception('Не найден элемент br_open')
                # код с переходом на следующий элемент
                text_log.insert(END, f"Номер {number}, не найден 'файлик' (D)\n")
                break
            # pause_check()
            progress_bar['value'] += 1
            window.update()

            br_open.click()         # нажимаем на кнопку с файликом
            logging.info('Click on the img_file')
            html = browser.page_source          # берем html страницы
            logging.info('Take the html')
            #logging.info(html)
            list_excel = parser(html, number)            # передаем html в парсер и создаем список list_excel
            logging.info('start insert data into excel')
            # pause_check()
            # начинаем подстановку в Excel
            r = cell.row
            c = cell.column + 1
            for i in list_excel:
                if progress_bar['value'] > 100:
                    progress_bar['value'] = 0
                    window.update()
                # pause_check()
                progress_bar['value'] += 1
                window.update()

                logging.info(f'insert {i}')
                try:
                    ws.cell(row=r, column=c, value=i)
                except Exception:
                    logging.exception('ошибка заполнения ячейки')
                else:
                    logging.info('Найдена ячейка и заплнена')
                c += 1
                # pause_check()
            try:
                wb.save(folder)
            except Exception:
                logging.exception('ошибка сохранения книги')
            else:
                logging.info('Сохранение книги')

                progress_bar['value'] += 1
                window.update()
            # pause_check()

            browser.back()          # переходим назад на страницу поиска по номеру
            logging.info('The Browser go back')
            # pause_check()
    wb.close()
    logging.info('The end of the parsing')
    text_log.configure(state='disabled')
    btn_check.configure(state='normal')
    btn_start.configure(state='normal')
    showinfo('Уведомление', 'Сбор информации завершен')

    progress_bar['value'] = 100
    window.update()
    # pause_check()
    logging.info('поток завершен')


def get_number2(folder):
    progress_bar['value'] = 0
    logging.info('Успешно вошли в функцию get_number2, в потоке')
    # pause_check()
    try:
        wb = load_workbook(folder)
        logging.info(f'Книга успешно открыта по адресу {folder}')
    except Exception:
        showerror('Ошибка', 'Не удается подключиться к книге Excel')
        logging.exception(f'Не удается подключиться к книге по пути {folder}')
        return
    # pause_check()
    try:
        ws = wb['Главный_лист']
        logging.info('Лист Excel успешно открыт')
    except Exception:
        showerror('Ошибка', 'Не найден лист')
        logging.exception(f'Лист не найден')
        wb.close()
        return
    # pause_check()
    try:
        wb.save(folder)
    except PermissionError:
        showerror('Ошибка', f'Необходимо закрыть файл {folder}. Закройте файл и нажмите на кнопку заново')
        logging.exception('файл открыт')
        wb.close()
        return
    else:
        logging.info('Файл закрыт, идем дальше')
    try:
        br_filter_open = browser.find_element(By.XPATH, xpath_filter_open)      # находим кнопку фильтра
        logging.info("The filter is opening successfully")
    except Exception:
        showerror('Не найден фильтр')
        logging.exception("The filter isn't opening successfully")
        wb.close()
        return
    progress_bar['value'] += 1
    # pause_check()
    br_filter_open.click()          # кликаем на кнопку фильтра
    numbers_excel = ws['C']         # берем все ячейки C
    count = len(numbers_excel)      # считаем кол-во ячеек непустых
    # pause_check()
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3, max_row=count):
        for cell in col:
            if progress_bar['value'] > 100:
                progress_bar['value'] = 0
                window.update()
            # pause_check()
            number = cell.value
            logging.info(f"Take number {number} from worklist")
            if number is None:          # если в excel пустоя ячейка (отсутствует номер), исключаем None. Иначе ошибка TypeError в send_keys()
                logging.info(f"Take number number from worklist, number is {number}, break")
                break
            try:
                br_number = browser.find_element(By.XPATH, xpath_number)  # находим поле с вводом номера регистрации
                logging.info('Элемент br_number упешно найден')
            except Exception:
                showerror('Ошибка', 'Не удается найти элемент')
                logging.exception('Элемент br_number не найден')
                wb.close()
                return
            # pause_check()
            try:
                br_apply = browser.find_element(By.XPATH, xpath_apply)  # находим поле с кнопкой "применить"
                logging.info('Элемент br_apply упешно найден')
            except Exception:
                showerror('Ошибка', 'Не удается найти элемент')
                logging.exception('Элемент br_apply не найден')
                wb.close()
                return
            # pause_check()
            logging.info('Clear the input textarea')
            br_number.clear()
            try:
                br_number.send_keys(number)         # отправляем номер в поле регистрации
                logging.info("the number from excel is sended in the br_number complitely ")
            except Exception:
                logging.exception("the number from excel is not sended in br_number")
                wb.close()
                return
            # pause_check()
            try:
                br_apply.click()            # нажимаем на кнопку "применить"
                logging.info("Click on the button_apply is successful")
            except Exception:
                logging.exception("Click on the button_apply isn't successful")
                wb.close()
                return
            sleep(2)
            try:
                br_open = browser.find_element(By.XPATH, xpath_open)           # находим забавный файлик
                logging.info('Элемент br_open найден')
            except Exception:
                # showerror('Ошибка', 'Не найден элемент')
                logging.exception('Не найден элемент br_open')
                # код с переходом на следующий элемент
                text_log.insert(END, f"Номер {number}, не найден 'файлик' (D)\n")
                break
            # pause_check()
            progress_bar['value'] += 1
            br_open.click()         # нажимаем на кнопку с файликом
            logging.info('Click on the img_file')
            html = browser.page_source          # берем html страницы
            logging.info('Take the html')
            #logging.info(html)
            list_excel = parser2(html, number)            # передаем html в парсер и создаем список list_excel
            logging.info('start insert data into excel')
            # pause_check()

            # начинаем подстановку в Excel
            r = cell.row
            c = cell.column + 2
            for i in list_excel:
                if progress_bar['value'] > 100:
                    progress_bar['value'] = 0
                    window.update()
                logging.info(f'insert {i}')
                # pause_check()
                progress_bar['value'] += 1
                if i == "":
                    c += 1
                    break
                try:
                    ws.cell(row=r, column=c, value=i)
                except Exception:
                    logging.exception('ошибка заполнения ячейки')
                else:
                    logging.info('Найдена ячейка и заполнена')
                c += 1
            try:
                wb.save(folder)
            except PermissionError:
                showerror('Ошибка', f'Необходимо закрыть файл {folder}')
                logging.exception('ошибка сохранения книги')
            else:
                logging.info('Сохранение книги')
            browser.back()          # переходим назад на страницу поиска по номеру
            logging.info('The Browser go back')
            # pause_check()
    wb.close()
    logging.info('The end of the parsing')
    text_log.configure(state='disabled')
    btn_start.configure(state='normal')
    btn_check.configure(state='normal')
    # pause_check()
    progress_bar['value'] = 100
    showinfo('Уведомление', 'Проверка завершена')
    logging.info('поток завершен')

def parser(html, number):
    # pause_check()
    progress_bar['value'] += 1
    window.update()

    logging.info('In parser function')
    list_excel = []
    try:
        soup = BeautifulSoup(html, 'html.parser')
        soup2 = BeautifulSoup(html,'lxml')
        logging.info("The soup is creating successful")
    except Exception:
        logging.exception("Problem with creating the soup")
        return
    # pause_check()

    # парсинг даты регистрации D
    pr_date_registration = soup.find_all('div', class_=class_date_registration)
    logging.info('parsing date of registration')
    if len(pr_date_registration) == 0:
        text_log.insert(END, f"Номер {number}, не найдена дата регистрации (D)\n")
        list_excel.append("")
    else:
        span_reg = pr_date_registration[0]
        date_reg = span_reg.find_next('span').string
        list_excel.append(date_reg)
        logging.info(f'Append to the list the date of reg. {date_reg}')

    # pause_check()
    # парсинг даты завершения работ E
    pr_date_end = soup.find_all('div', class_=class_date_end)
    logging.info('parsing date of ending')
    if len(pr_date_end) == 0:
        text_log.insert(END, f"Номер {number}, не найдена дата окончания работ (E)\n")
        list_excel.append("")
    else:
        span_end = pr_date_end[0]
        date_end = span_end.find_next('span').string
        list_excel.append(date_end)
        logging.info(f'append in list date of end {date_end}')
    # pause_check()

    # парсинг уведомления о завершении работ F
    pr_notify_end = soup.find_all('div', class_='field-name-field-fgpn-notify-end-rel')
    if len(pr_notify_end) == 0:
        text_log.insert(END, f"Номер {number}, не найдено уведомление о завершении работ (F)\n")
        list_excel.append("")
    else:
        a = pr_notify_end[0]
        notify_end = a.find_next('a').string
        m = re.search('..-..-....-......', notify_end)
        if m is None:
            list_excel.append("")
            text_log.insert(END, f"Номер {number}, не найдено уведомление о завершении работ (F)\n")
            logging.info('не найдена подстрока в строке')
        else:
            list_excel.append(m.group())
        logging.info(f'append in list the {m.group()}')
    # pause_check()

    # парсинг лицензиат G
    pr_license = soup.find_all('div', class_=class_license)
    logging.info('parsing pr_license')
    if len(pr_license) == 0:
        text_log.insert(END, f"Номер {number}, не найден лицензиат (G)\n")
        list_excel.append("")
    else:
        a_lic = pr_license[0]
        licen = a_lic.find_next('a').string
        list_excel.append(licen)
        logging.info(f'append the {licen}')
    # pause_check()

    # парсинг номер лицензии H
    pr_number_l = soup.find_all('div', class_=class_number_l)
    logging.info('parsing pr_number_l')
    if len(pr_number_l) == 0:
        text_log.insert(END, f"Номер {number}, не найден номер лицензии (H)\n")
        list_excel.append("")
    else:
        div_number_l = pr_number_l[0]
        number_l = div_number_l.find_next('div', class_='field-item').string
        list_excel.append(number_l)
        logging.info(f'append in list the {number_l}')

    progress_bar['value'] += 1
    window.update()
    # pause_check()

    #парсинг места осуществления деятельности I
    try:
        br_open_a = browser.find_element(By.XPATH, xpath_open_a)
    except Exception:
        # showerror('Ошибка', 'Нет перехода по ссылке Лицензия')
        logging.exception('Ошибка перехода по ссылке')
        text_log.insert(END, f"Номер {number}, не найдено место осуществления деятельности (проблема с переходом по ссылке Лицензия) (I)\n")
        list_excel.append("")
    else:
        br_open_a.click()
        logging.info('Переход по ссылке для парсинга I')
        html2 = browser.page_source
        # logging.info(html2)
        logging.info('Берем html2')
        soup3 = BeautifulSoup(html2, 'lxml')
        pr_places_work = soup3.find_all('div', class_='field-name-field-fgpn-places')
        browser.back()
        if len(pr_places_work) == 0:
            text_log.insert(END, f"Номер {number}, не найдено место осуществления деятельности (I)\n")
            list_excel.append("")
        else:
            a = pr_places_work[0].find_next('a').text
            l = a.replace("Адрес: ","")
            list_excel.append(l)
            logging.info(f'append in list the {l}')
    # pause_check()

    # парсинг адреса выполнения работ J
    pr_address_work = soup2.find('div', class_='field-name-field-gl-adresses')
    logging.info('parsing pr_address_work')
    if len(pr_address_work) == 0:
        text_log.insert(END, f"Номер {number}, не найдены адреса выполнения работ (J)\n")
        list_excel.append("")
    else:
        div_list = pr_address_work.find_all('div', class_='field-item')
        addresses_work = ""
        for i in div_list:
            addresses_work += i.text + '\n'
        list_excel.append(addresses_work)
        logging.info(f'append in list the {addresses_work}')
    # pause_check()

    # парсинг даты заявления K
    pr_receive_date = soup.find_all('div', class_='field-name-field-gl-receive-date')
    logging.info('parsing pr_receive_date')
    if len(pr_receive_date) == 0:
        text_log.insert(END, f"Номер {number}, не найдена дата заявления (K)\n")
        list_excel.append("не найдено")
    else:
        span_receive_date = pr_receive_date[0]
        receive_date = span_receive_date.find_next('span').string
        list_excel.append(receive_date)
        logging.info(f'append in list the {receive_date}')
    # pause_check()

    # парсинг даты договора L
    pr_data_deal = soup.find_all('div', class_='field-name-field-fgpn-notify-contract--date')
    logging.info('parsing pr_data_deal')
    if len(pr_data_deal) == 0:
        text_log.insert(END, f"Номер {number}, не найдена дата договора (L)\n")
        list_excel.append("")
    else:
        div_data_deal = pr_data_deal[0]
        data_deal = div_data_deal.find_next('div', class_='field-item even').string
        list_excel.append(data_deal)
        logging.info(f'append in list the {data_deal}')

    progress_bar['value'] += 1
    window.update()
    # pause_check()

    # парсинг номер договора M
    pr_number_deal = soup.find_all('div', class_='field-name-field-fgpn-notify-contract--number')
    logging.info('parsing pr_number_deal')
    if len(pr_number_deal) == 0:
        text_log.insert(END, f"Номер {number}, не найден номер договора (M)\n")
        list_excel.append("")
    else:
        div_number_deal = pr_number_deal[0]
        number_deal = div_number_deal.find_next('div', class_='field-item even').string
        list_excel.append(number_deal)
        logging.info(f'append in list the {number_deal}')
    # pause_check()

    # парсинг заказчика работ N
    pr_customer = soup.find_all('div', class_='field-name-field-fgpn-notify-contract--customer')
    logging.info('parsing pr_customer')
    if len(pr_customer) == 0:
        text_log.insert(END, f"Номер {number}, не найден заказчик работ (N)\n")
        list_excel.append("")
    else:
        div_customer = pr_customer[0]
        customer = div_customer.find_next('div', class_='field-item even').string
        list_excel.append(customer)
        logging.info(f'append in the list the {customer}')
    # pause_check()

    # парсинг инн O
    pr_inn = soup.find_all('div', class_='field-name-field-fgpn-notify-contract--inn')
    logging.info('parsing pr_inn')
    if len(pr_inn) == 0:
        text_log.insert(END, f"Номер {number}, не найден инн (О)\n")
        list_excel.append("")
    else:
        div = pr_inn[0]
        inn = div.find_next('div', class_='field-item even').string
        list_excel.append(inn)
        logging.info(f'append in the list the {inn}')
    # pause_check()

    # парсинг объекта P
    pr_object = soup.find_all('div', class_='field-name-field-fgpn-object-name')
    if len(pr_object) == 0:
        text_log.insert(END, f"Номер {number}, не найдено имя объекта (P)\n")
        list_excel.append("")
    else:
        div = pr_object[0]
        object_name = div.find_next('div', class_='field-item even').string
        list_excel.append(object_name)
        logging.info(f'append in the list the {object_name}')
    # pause_check()

    # парсинг вид работы Q
    pr_kind_work = soup.find_all('div', class_='field-name-field-fgpn-notify-kind')
    if len(pr_kind_work) == 0:
        text_log.insert(END, f"Номер {number}, не найден вид работы (Q)\n")
        list_excel.append("")
    else:
        div = pr_kind_work[0]
        kind_work = div.find_next('div', class_='field-item even').string
        list_excel.append(kind_work)
        logging.info(f'append in the list the {kind_work}')
    # pause_check()

    # парсинг номер проекта R
    pr_object_number = soup.find_all('div', class_='field-name-field-fgpn-notify-project--number')
    if len(pr_object_number) == 0:
        text_log.insert(END, f"Номер {number}, не найден номер проекта (R)\n")
        list_excel.append("")
    else:
        div = pr_object_number[0]
        object_number = div.find_next('div', class_='field-item even').string
        list_excel.append(object_number)
        logging.info(f'append in the list the {object_number}')

    progress_bar['value'] += 1
    window.update()
    # pause_check()

    # парсинг дата проекта S
    pr_project_data = soup.find_all('div', class_='field-name-field-fgpn-notify-project--date')
    if len(pr_project_data) == 0:
        text_log.insert(END, f"Номер {number}, не найдена дата проекта (S)\n")
        list_excel.append("")
    else:
        div = pr_project_data[0]
        project_data = div.find_next('div', class_='field-item even').string
        list_excel.append(project_data)
        logging.info(f'append in the list the {project_data}')
    # pause_check()

    # парсинг фамилия проектировщика T
    pr_author_surname = soup.find_all('div', class_='field-name-field-fgpn-notify-project-author--f')
    if len(pr_author_surname) == 0:
        text_log.insert(END, f"Номер {number}, не найдена фамилия проектировщика (T)\n")
        list_excel.append("")
    else:
        div = pr_author_surname[0]
        author_surname = div.find_next('div', class_='field-item even').string
        list_excel.append(author_surname)
        logging.info(f'append in the list the {author_surname}')
    # pause_check()

    # парсинг имени проектировщика U
    pr_author_name = soup.find_all('div', class_='field-name-field-fgpn-notify-project-author--i')
    if len(pr_author_name) == 0:
        text_log.insert(END, f"Номер {number}, не найдено имя проектировщика (U)\n")
        list_excel.append("")
    else:
        div = pr_author_name[0]
        author_name = div.find_next('div', class_='field-item even').string
        list_excel.append(author_name)
        logging.info(f'append in the list the {author_name}')

    progress_bar['value'] += 1
    window.update()
    # pause_check()

    # парсинг отчества проектировщика V
    pr_author_ot = soup.find_all('div', class_='field-name-field-fgpn-notify-project-author--o')
    if len(pr_author_ot) == 0:
        text_log.insert(END, f"Номер {number}, не найдено отчество проектировщика (V)\n")
        list_excel.append("")
    else:
        div = pr_author_ot[0]
        author_ot = div.find_next('div', class_='field-item even').string
        list_excel.append(author_ot)
        logging.info(f'append in the list the {author_ot}')

    # pause_check()

    # парсинг номер аттестата W
    pr_number_diplom = soup.find_all('div', class_='field-name-field-fgpn-notify-project-author--cert-number')
    if len(pr_number_diplom) == 0:
        text_log.insert(END, f"Номер {number}, не найден номер аттестата (W)\n")
        list_excel.append("")
    else:
        div = pr_number_diplom[0]
        number_diplom = div.find_next('div', class_='field-item even').string
        list_excel.append(number_diplom)
        logging.info(f'append in the list the {number_diplom}')
    # pause_check()

    # парсинг даты аттестата X
    pr_data_diplom = soup.find_all('div', class_='field-name-field-fgpn-notify-project-author--cert-date')
    if len(pr_data_diplom) == 0:
        text_log.insert(END, f"Номер {number}, не найдена дата аттестат (X)\n")
        list_excel.append("")
    else:
        div = pr_data_diplom[0]
        data_diplom = div.find_next('div', class_='field-item even').string
        list_excel.append(data_diplom)
        logging.info(f'append in the list the {data_diplom}')

    # pause_check()

    # парсинг ответственного фамилия Y
    pr_gl_employee = soup.find_all('div', class_='field-name-field-gl-employee--f')
    if len(pr_gl_employee) == 0:
        text_log.insert(END, f"Номер {number}, не найдена фамилия ответственного (Y)\n")
        list_excel.append("")
    else:
        div = pr_gl_employee[0]
        gl_employee = div.find_next('div', class_='field-item even').string
        list_excel.append(gl_employee)
        logging.info(f'append in the list the {gl_employee}')
    # pause_check()

    # парсинг ответственного имя Z
    pr_gl_employee_name = soup.find_all('div', class_='field-name-field-gl-employee--i')
    if len(pr_gl_employee_name) == 0:
        text_log.insert(END, f"Номер {number}, не найдено имя ответственного (Z)\n")
        list_excel.append("")
    else:
        div = pr_gl_employee_name[0]
        gl_employee_name = div.find_next('div', class_='field-item even').string
        list_excel.append(gl_employee_name)
        logging.info(f'append in the list the {gl_employee_name}')
    # pause_check()

    # парсинг ответственного отчество АА
    pr_gl_employee_ot = soup.find_all('div', class_='field-name-field-gl-employee--o')
    if len(pr_gl_employee_ot) == 0:
        text_log.insert(END, f"Номер {number}, не найдено отчество ответственного (АА)\n")
        list_excel.append("")
    else:
        div = pr_gl_employee_ot[0]
        gl_employee_ot = div.find_next('div', class_='field-item even').string
        list_excel.append(gl_employee_ot)
        logging.info(f'append in the list the {gl_employee_ot}')
    # pause_check()

    # парсинг ответственного снилса АB
    pr_gl_employee_snils = soup.find_all('div', class_='field-name-field-gl-employee--snils')
    if len(pr_gl_employee_snils) == 0:
        text_log.insert(END, f"Номер {number}, не найден СНИЛС ответственного (АВ)\n")
        list_excel.append("")
    else:
        div = pr_gl_employee_snils[0]
        gl_employee_snils = div.find_next('div', class_='field-item even').string
        list_excel.append(gl_employee_snils)
        logging.info(f'append in the list the {gl_employee_snils}')

    progress_bar['value'] += 1
    window.update()
    # pause_check()

    # парсинг оборудования, начиная с AC
    data = []
    table = soup2.find('table', class_='tableheader-processed')
    if len(table) == 0:
        text_log.insert(END, f"Номер {number}, не найдено оборудование (ячейки, начиная с АC)\n")
        list_excel.append("")
    else:
        table_body = table.find('tbody')
        rows = table_body.find_all('tr')
        for row in rows:
            cols = row.find_all('td')
            cols = [ele.text.strip() for ele in cols]
            data.append([ele for ele in cols if ele])

        for i in data:
            for j in range(2):
                list_excel.append(i[j])
                logging.info(f'append in the list the {i[j]}')

    return list_excel

def parser2(html, number):
    # pause_check()
    progress_bar['value'] += 10
    window.update()
    logging.info('In parser function')
    list_excel = []
    try:
        soup = BeautifulSoup(html, 'html.parser')
        # soup2 = BeautifulSoup(html, 'lxml')
        logging.info("The soup is creating successful")
    except Exception:
        logging.exception("Problem with creating the soup")
        return
    # pause_check()

    # парсинг даты завершения работ E
    pr_date_end = soup.find_all('div', class_=class_date_end)
    logging.info('parsing date of ending')
    if len(pr_date_end) == 0:
        text_log.insert(END, f"Номер {number}, не найдена дата окончания работ (E)\n")
        list_excel.append("")
    else:
        span_end = pr_date_end[0]
        date_end = span_end.find_next('span').string
        list_excel.append(date_end)
        logging.info(f'append in list date of end {date_end}')
    # pause_check()
    progress_bar['value'] += 10
    window.update()

    # парсинг уведомления о завершении работ F
    pr_notify_end = soup.find_all('div', class_='field-name-field-fgpn-notify-end-rel')
    if len(pr_notify_end) == 0:
        text_log.insert(END, f"Номер {number}, не найдено уведомление о завершении работ (F)\n")
        list_excel.append("")
    else:
        a = pr_notify_end[0]
        notify_end = a.find_next('a').string
        m = re.search('..-..-....-......', notify_end)
        # pause_check()
        if m is None:
            list_excel.append("")
            text_log.insert(END, f"Номер {number}, не найдено уведомление о завершении работ (F)\n")
            logging.info('не найдена подстрока в строке')
        else:
            list_excel.append(m.group())
        logging.info(f'append in list the {m.group()}')

    return list_excel

def start():
    btn_check.configure(state='disabled')
    text_log.configure(state='normal')
    folder_xl = txt_folder_xl.get()
    get_number(folder_xl)
    # logging.info('Создаем поток')
    # t1 = Thread(target=get_number, args=folder_xl, daemon=True)
    # logging.info('старт потока')
    # t1.start()

    # btn_pause.configure(state='normal')

def check():
    btn_start.configure(state='disabled')
    text_log.configure(state='normal')
    folder_xl = txt_folder_xl.get()
    # logging.info('Создаем поток')
    # t1 = Thread(target=get_number2, args=folder_xl, daemon=True)
    # logging.info('старт потока')
    # t1.start()
    get_number2(folder_xl)
    # btn_pause.configure(state='normal')

# def pause():
#     global flag_pause, k
#     logging.info('нажата кнопка пауза')
#     if k % 2 == 0:
#         btn_pause.configure(state='disabled')
#         k += 1
#         flag_pause = True
#     elif k % 2 == 1:
#         btn_pause.configure(text='Пауза')
#         k += 1
#         flag_pause = False

def analiz_click_t():
    down = lbl_range_down.get()
    up = lbl_range_up.get()
    folder_xl = txt_folder_xl.get()
    if folder_xl == '':
        showerror('Ошибка','Выберите файл')
        return
    try:
        down = int(down)
        up = int(up)
    except ValueError:
        showerror('Ошибка', 'Введены не целые числа')
        return
    else:
        if down > up:
            showerror('Ошибка', 'Введите числа по возрастанию')
            return
    btn_analiz.configure(state='disabled')
    btn_check.configure(state='disabled')
    btn_start.configure(state='disabled')
    t1 = Thread(target=analiz_click, args=(down, up, folder_xl), daemon=True)
    t1.start()
    t2 = Thread(target=check_progres, args=(), daemon=True)
    t2.start()

def check_progres():
    while len(threading.enumerate()) == 3:
        progress_bar.start(100)
    progress_bar.stop()
    progress_bar['value'] = 100


def analiz_click(down, up, folder_xl):
    analiz_snils.main_an(down, up, folder_xl)
    analiz_place.main_an2(down, up, folder_xl)
    # analiz_serial_number.main_an(down, up, folder_xl)
    btn_check.configure(state='normal')
    btn_start.configure(state='normal')
    btn_analiz.configure(state='normal')

def open_folder():
    file_name = fd.askopenfilename()
    if file_name[-5:] in (".xlsx",".xlsm") or file_name[-4:] == ".xls":
        pass
    else:
        showerror('Ошибка','Выберите файл .xlsx')
        return
    txt_folder_xl.insert(0, file_name)

window = Tk()
window.title('Поиск информации')
window.geometry('450x450')
lbl_login = Label(window, text="Логин:")
lbl_password = Label(window, text="Пароль:")
lbl_folder = Label(window, text='Расположение файла Excel:')
txt_login = Entry(window, width=20)
txt_password = Entry(window, width=20)
txt_folder_xl = Entry(window, width=20, state='normal')
btn_entry = Button(window, text='Войти', width=17, command=entry)
btn_start = Button(window, text='Начать заполнение', state='disabled', command=start)
btn_check = Button(window, text='Проверить ячейки', state='disabled', command=check)
# btn_pause = Button(window, text='Пауза', state='disabled', command=pause)
btn_analiz = Button(window, text="Анализ", command=analiz_click_t)
progress_bar = ttk.Progressbar(window, orient="horizontal",
                               mode="determinate", maximum=100, value=0)
btn_open = Button(window, text="Открыть", command=open_folder)

text_log = scrolledtext.ScrolledText(window, width=38, height=10, wrap=tkinter.WORD, state='disabled')

lbl_login.grid(column=0, row=0, sticky=E)
txt_login.grid(column=1, row=0)
progress_bar.grid(column=2, row=0)
btn_open.grid(column=2, row=3, sticky=W)
window.update()
progress_bar['value'] = 0
window.update()
txt_down = Label(window, text="Проверка ячеек: от")
txt_up = Label(window, text="до")
lbl_range_down = Entry(window, width=5)
lbl_range_up = Entry(window, width=5)

lbl_password.grid(column=0, row=1, sticky=E)
txt_password.grid(column=1, row=1)
btn_entry.grid(column=1, row=2, pady=10)
lbl_folder.grid(column=0, row=3, sticky=E)
txt_folder_xl.grid(column=1, row=3)
btn_start.grid(column=0, row=4, pady=10)
btn_check.grid(column=1, row=4, pady=10)
# btn_pause.grid(column=2, row=4, pady=10)
text_log.grid(column=0, row=5, pady=10, padx=10, columnspan=2)
txt_down.grid(column=0, row=6, sticky=E)
lbl_range_down.grid(column=1, row=6, sticky=W)
txt_up.grid(column=0, row=7, sticky=E)
lbl_range_up.grid(column=1, row=7, sticky=W)
btn_analiz.grid(column=2, row=7, sticky=W)
# window.iconbitmap('3-search-cat_icon-icons.com_76679.ico')
window.mainloop()
