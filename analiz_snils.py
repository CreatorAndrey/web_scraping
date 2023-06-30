import copy
from tkinter.messagebox import showerror, showinfo
from copy import copy as cop
from openpyxl import load_workbook
# from main_mcs import progress_bar, window
import tkinter.ttk as ttk

k = 0


''' цикл собирает все СНИЛС в словарь, ключ-СНИЛС, значение-номера столбцов в excel {'1111': [1, 4, 6]} '''
def get_snils(folder):
    global ws, wb, ws_an, max_row1
    try:
        wb = load_workbook(folder)
        ws = wb['Главный_лист']
        ws_an = wb['Анализ']
    except Exception:
        showerror('Ошибка', 'Проблема открытия книги')
        exit()
    # progress_bar['value'] = 5
    # window.update()
    max_row1 = ws_an.max_row
    # print(max_row1)
    numbers = ws['AB']
    count = len(numbers)
    # print(count)
    dict_snils = {}          # словарь со всеми СНИЛС
    # progress_bar['value'] = 10
    # window.update()
    for col in ws.iter_cols(min_row=2, min_col=28, max_col=28, max_row=count):
        for cell in col:
            # progress_bar['value'] += 1
            # window.update()
            # if progress_bar['value'] > 100:
            #     progress_bar['value'] = 0
            snils = cell.value
            # print(snils)
            if snils is None:  # если в excel пустоя ячейка (отсутствует номер), исключаем None. Иначе ошибка TypeError в send_keys()
                continue
            if snils in dict_snils.keys():
                indexes = dict_snils.get(snils)
                indexes.append(cell.row)
                dict_snils[snils] = indexes
            else:
                dict_snils[snils] = [cell.row]
    return dict_snils

def copy_cell(src_sheet, src_row, src_col, tgt_sheet,
              tgt_row, tgt_col, copy_style=True):
    cell = src_sheet.cell(src_row, src_col)
    new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
    if cell.has_style and copy_style:
        new_cell._style = cop(cell._style)

def delete_F(indexes, snils, dict_range, folder):
    global k
    list_val = []
    ''' удаляем из списка индексов индексы с существующим F'''
    indexes2 = copy.deepcopy(indexes)
    for i in indexes:
        # progress_bar['value'] += 1
        # window.update()
        # if progress_bar['value'] > 100:
        #     progress_bar['value'] = 0
        if ws[f'F{i}'].value:
            indexes2.remove(i)
    # print(indexes2)
    indexes.clear()

    ''' проверяем совпадения H'''
    dict_H = {}
    for i in indexes2:
        # progress_bar['value'] += 1
        # window.update()
        # if progress_bar['value'] > 100:
        #     progress_bar['value'] = 0
        val_H = ws[f'H{i}']
        if val_H.value is None:
            continue
        if val_H.value in dict_H.keys():
            val = dict_H.get(val_H.value)
            val.append(val_H.row)
            dict_H[val_H.value] = val
        else:
            dict_H[val_H.value] = [val_H.row]
    # print(dict_H)

    for key, value in dict_H.items():
        if (len(value) > 1) & (snils in dict_range):
            for i in value:
                for j in range(1, ws.max_column + 1):
                    # progress_bar['value'] += 1
                    # window.update()
                    # if progress_bar['value'] > 100:
                    #     progress_bar['value'] = 0
                    cell_obj = ws.cell(row=i, column=j)
                    # print(cell_obj.value, end=" ")

                    cell_1 = cell_obj.column

                    copy_cell(ws, i, cell_1, ws_an, max_row1 + k, cell_1)
                # print()
                print(i)
                k += 1
                # print(k)
                # wb.save(folder)










def main_an(down, up, folder):
    try:
        dict_snils = get_snils(folder)
        dict_snils_range = set()
        try:
            wb.save(folder)
        except PermissionError:
            showerror('Ошибка', f'Необходимо закрыть файл {folder}. Закройте файл и нажмите на кнопку заново заново')
            # logging.exception('файл открыт')
            wb.close()
            return

        for col in ws.iter_cols(min_row=down, min_col=28, max_col=28, max_row=up):
            for cell in col:
                # progress_bar['value'] += 1
                # window.update()
                # if progress_bar['value'] > 100:
                #     progress_bar['value'] = 0
                snils = cell.value
                if snils is None:
                    continue
                # print(snils)
                dict_snils_range.add(snils)

        for key, value in dict_snils.items():
            if len(value) > 1:
                delete_F(value, key, dict_snils_range, folder)

        wb.save(folder)
        wb.close()
    except:
        wb.save(folder)
        wb.close()



# list_snils_iter = copy.deepcopy(list_snils)
# for i in list_snils_iter:
#     list_snils.pop(0)
#     if i in list_snils:
