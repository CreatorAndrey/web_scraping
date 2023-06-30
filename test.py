from openpyxl import load_workbook

wb = load_workbook('D:\Работа\Авито парсер локальный сайт\ТЕКУЩИЙ.xlsx')
ws_an = wb['Анализ']
max_row1 = ws_an.max_row
print(max_row1)