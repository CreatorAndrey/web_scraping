from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from tkinter.messagebox import showerror, showinfo
import main_mcs
from time import sleep


xpath_apply = '//*[@id="edit-submit-fgpn-license-notification-record"]'         # кнопка применить
xpath_number = '//*[@id="edit-field-gl-registry-num-value"]'           # ссылка на поле с вводом номера регистрации
xpath_open = '//*[@id="block-system-main"]/div/div[2]/div/table/tbody/tr/td[12]/div/a/span'     # забавный файлик

def get_number(folder):
    try:
        wb = load_workbook(folder)
    except:
        showerror('Ошибка', 'Не удается подключиться к книге Excel')
        return
    try:
        ws = wb['Главный_лист']
    except:
        showerror('Ошибка', 'Не найден лист')
        return
    try:
        br_number = main_mcs.browser.find_element(By.XPATH, xpath_number)
        br_apply = main_mcs.browser.find_element(By.XPATH, xpath_apply)
    except:
        showerror('Ошибка', 'Не удается найти элемент по xpath')
        return
    numbers_excel = ws['C']
    count = len(numbers_excel)
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3, max_row=count, values_only=True):
        for cell in col:
            number = cell
            br_number.send_keys(number)
            br_apply.click()
            sleep(2)
            try:
                br_open = main_mcs.browser.find_element(By.XPATH, xpath_open)           # нажимаем на забавный файлик
            except:
                showerror('Ошибка', 'Не найден элемент')
                #код с переходом на следующий элемент
                return
            br_open.click()
            showinfo('Ок','поиск по номеру завершен')



